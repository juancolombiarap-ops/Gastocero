# -*- coding: utf-8 -*-
"""Arma el Excel de insumo del Reporte de Competencia para un periodo nuevo.

El layout es el que espera mapa.py: columna A con nombres, E-J con precios,
fila 'HOTEL' declarando los canales, filas de sección (ALOJAMIENTO/PAQUETERÍA),
de grupo (HOTELES DECAMERON / COMPETENCIA) y de destino (ADZ, CTG, ...).
Las fechas del periodo van en T2/U2 de la hoja TABLA, que es de donde
informe.py las lee: no se escriben a mano en ninguna otra parte.

Uso: py -3 generar_plantilla.py 2026-10-03 2026-10-08 [SALIDA.xlsx]
"""
import datetime as dt
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

CANALES = ['V. FALABELLA', 'DESPEGAR', 'COCHA', 'EXPEDIA']
AEROLINEAS = ['LATAM', 'AVIANCA', 'COPA', 'SKY', 'JETSMART', 'ARAJET']
RUTAS = ['SCL-ADZ', 'SCL-CTG', 'SCL-TBP', 'SCL-PUJ', 'SCL-CUN', 'SCL-SMR', 'SCL-PTY']

# Nómina cotizada en el informe 10-15 Oct 2026 V.2, por destino.
# (destino, [(grupo, hotel), ...]) — el orden es el del informe.
DESTINOS = [
    ('ADZ', [
        ('DECAMERON', 'DECAMERON ISLEÑO'),
        ('DECAMERON', 'DECAMERON SAN LUIS'),
        ('DECAMERON', 'DECAMERON MARAZUL'),
        ('DECAMERON', 'DECAMERON MARYLAND'),
        ('DECAMERON', 'DECAMERON LOS DELFINES'),
        ('COMPETENCIA', 'SOL CARIBE SAN ANDRES'),
        ('COMPETENCIA', 'SOL CARIBE CAMPO'),
        ('COMPETENCIA', 'EL DORADO'),
        ('COMPETENCIA', 'GRAND SIRENIS SAN ANDRES'),
    ]),
    ('CTG', [
        ('DECAMERON', 'DECAMERON CARTAGENA'),
        ('DECAMERON', 'DECAMERON BARU'),
        ('COMPETENCIA', 'DORADO PLAZA BOCAGRANDE'),
        ('COMPETENCIA', 'HOTEL CARTAGENA DUBAI'),
        ('COMPETENCIA', 'HOTEL LA GRAN VIA'),
        ('COMPETENCIA', 'CARTAGENA PLAZA'),
    ]),
    ('TBP', [
        ('DECAMERON', 'ROYAL DECAMERON PUNTA SAL'),
    ]),
    ('PTY', [
        ('DECAMERON', 'GRAND DECAMERON PANAMA (7.5)'),
        ('DECAMERON', 'GRAND DECAMERON PANAMA (7.8)'),
        ('COMPETENCIA', 'RIU PLAYA BLANCA ALL INCLUSIVE'),
        ('COMPETENCIA', 'GRAN EVENIA BIJAO'),
        ('COMPETENCIA', 'DREAMS PLAYA BONITA PANAMA'),
    ]),
    ('SMR', [
        ('DECAMERON', 'DECAMERON GALEON'),
        ('COMPETENCIA', 'HOTEL PORTO HORIZONTE'),
        ('COMPETENCIA', 'IROTAMA DEL SOL'),
    ]),
    ('MEXICO', [
        ('DECAMERON', 'GRAND DECAMERON COMPLEX'),
        ('DECAMERON', 'GRAND DECAMERON LOS CABOS'),
    ]),
    ('PUJ', [
        ('COMPETENCIA', 'BARCELO BAVARO PALACE'),
        ('COMPETENCIA', 'SERENADE CARIBE CLUB FAMILY'),
        ('COMPETENCIA', 'GRAND SIRENIS PUNTA CANA'),
        ('COMPETENCIA', 'VIK HOTEL CAYENA'),
    ]),
    ('CUN', [
        ('COMPETENCIA', 'RIU DUNAMAR'),
        ('COMPETENCIA', 'FLAMINGO CANCUN'),
        ('COMPETENCIA', 'SUNSET MARINA RESORT'),
        ('COMPETENCIA', 'DREAMS RIVIERA'),
        ('COMPETENCIA', 'HYATT VIVID CANCUN'),
    ]),
    ('CURAZAO', [
        ('COMPETENCIA', 'ZOËTRY CURAÇAO RESORT & SPA'),
        ('COMPETENCIA', 'MANGROVE BEACH CORENDON CURACAO ALL-INCLUSIVE RESORT, CURIO BY HILTON'),
        ('COMPETENCIA', 'DREAMS CURACAO'),
    ]),
]

AZUL = 'FF003A70'
CELESTE = 'FFE3F1F8'
GRIS = 'FFF6F9FC'


def _titulo(ws, row, texto, ancho=10):
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, color='FFFFFFFF', size=12)
    c.fill = PatternFill('solid', fgColor=AZUL)
    for col in range(2, ancho + 1):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=AZUL)


def _cabecera(ws, row):
    c = ws.cell(row=row, column=1, value='HOTEL')
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor=CELESTE)
    ws.cell(row=row, column=2, value='CHECK IN').font = Font(bold=True)
    ws.cell(row=row, column=3, value='CHECK OUT').font = Font(bold=True)
    for i, canal in enumerate(CANALES):
        cc = ws.cell(row=row, column=5 + i, value=canal)
        cc.font = Font(bold=True)
        cc.fill = PatternFill('solid', fgColor=CELESTE)
        cc.alignment = Alignment(horizontal='center')


def _bloque(ws, row, seccion, ini, fin):
    """Escribe una sección completa. Compacto a propósito: mapa.py solo escanea
    las primeras 135 filas, así que alojamiento y paquetería tienen que caber
    juntas ahí. Por eso la fila HOTEL va una sola vez por sección (mapa.py
    arrastra los canales declarados hasta la siguiente cabecera)."""
    _titulo(ws, row, seccion)
    row += 1
    _cabecera(ws, row)
    row += 1
    for dest, hoteles in DESTINOS:
        ws.cell(row=row, column=1, value=dest).font = Font(bold=True, color=AZUL)
        row += 1
        grupo_actual = None
        for grupo, hotel in hoteles:
            if grupo != grupo_actual:
                etq = 'HOTELES DECAMERON' if grupo == 'DECAMERON' else 'COMPETENCIA'
                ws.cell(row=row, column=1, value=etq).font = Font(bold=True, italic=True)
                grupo_actual = grupo
                row += 1
            ws.cell(row=row, column=1, value=hotel)
            ws.cell(row=row, column=2, value=ini)
            ws.cell(row=row, column=3, value=fin)
            for i in range(len(CANALES)):
                celda = ws.cell(row=row, column=5 + i)
                celda.fill = PatternFill('solid', fgColor=GRIS)
                celda.number_format = '#,##0'
            row += 1
    return row


def construir(ini, fin, salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TABLA'

    # Periodo cotizado: informe.py lo lee de T2/U2 y de ahí sale todo el rotulado.
    ws['S2'] = 'PERIODO'
    ws['S2'].font = Font(bold=True)
    ws['T2'] = ini
    ws['U2'] = fin
    for celda in ('T2', 'U2'):
        ws[celda].number_format = 'DD/MM/YYYY'

    row = _bloque(ws, 4, 'ALOJAMIENTO', ini, fin)
    row += 1
    _bloque(ws, row, 'PAQUETERIA', ini, fin)

    ws.column_dimensions['A'].width = 52
    for col in 'BC':
        ws.column_dimensions[col].width = 12
    for col in 'EFGHIJ':
        ws.column_dimensions[col].width = 14

    wa = wb.create_sheet('AEREO')
    wa['B2'] = 'VUELO'
    wa['B2'].font = Font(bold=True)
    for i, aero in enumerate(AEROLINEAS):
        c = wa.cell(row=2, column=6 + i, value=aero)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor=CELESTE)
    for j, ruta in enumerate(RUTAS):
        wa.cell(row=3 + j, column=2, value=ruta)
        for i in range(len(AEROLINEAS)):
            celda = wa.cell(row=3 + j, column=6 + i)
            celda.fill = PatternFill('solid', fgColor=GRIS)
            celda.number_format = '#,##0'
    wa.column_dimensions['B'].width = 14
    for col in 'FGHIJK':
        wa.column_dimensions[col].width = 12

    wi = wb.create_sheet('INSTRUCCIONES')
    for i, linea in enumerate([
        'REPORTE DE COMPETENCIA — planilla de insumo',
        '',
        f'Periodo cotizado: {ini:%d/%m/%Y} al {fin:%d/%m/%Y} (5 noches, 2 adultos).',
        'Las fechas viven en T2/U2 de la hoja TABLA. Cambiarlas ahí y solo ahí:',
        'informe.py rotula todo el PDF a partir de esas dos celdas.',
        '',
        'Qué se carga: total de la estadía en USD, 2 adultos, habitación Standard,',
        'tarifa más económica de esa categoría, All Inclusive donde aplique.',
        'Si el hotel no tiene Standard en esas fechas se cotiza la categoría',
        'siguiente y se declara en CORRECCION_CATEGORIA_POR_PERIODO de informe.py.',
        '',
        'Celda vacía = canal no cotizado (sale N/D en el informe).',
        'Cero o texto = cotizado sin disponibilidad ni tarifa (sale NA).',
        '',
        'Generar el informe:',
        '  py -3 informe.py "ESTE_ARCHIVO.xlsx" --version colombia',
        '  py -3 informe.py "ESTE_ARCHIVO.xlsx" --version interna \\',
        '        --dchile-json dchile_03-08_Oct_2026.json --dchile-fecha "29/08/2026 17:00"',
    ], start=2):
        wi.cell(row=i, column=2, value=linea)
    wi.column_dimensions['B'].width = 90

    wb.save(salida)
    return salida


if __name__ == '__main__':
    ini = dt.date.fromisoformat(sys.argv[1])
    fin = dt.date.fromisoformat(sys.argv[2])
    out = sys.argv[3] if len(sys.argv) > 3 else (
        f'Reporte_Competencia_{ini.day:02d}-{fin.day:02d}_{ini:%b}_{ini.year}_PLANTILLA.xlsx')
    print(construir(ini, fin, out))
